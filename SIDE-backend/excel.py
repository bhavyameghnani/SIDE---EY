#import library
from openpyxl import load_workbook
import pandas as pd


def read_excel(filename):
    #read file
    wb = load_workbook(filename)

    #access specific sheet
    ws = wb["Sheet1"]

    print(ws)
    print(ws.tables.items)
    mapping = {}

    for entry, data_boundary in ws.tables.items():
        #parse the data within the ref boundary
        
        data = ws[data_boundary]
        print(data)
        #extract the data 
        #the inner list comprehension gets the values for each cell in the table
        content = [[cell.value for cell in ent] 
                for ent in data
            ]
        header = content[0]
        
        #the contents ... excluding the header
        rest = content[1:]
        
        #create dataframe with the column names
        #and pair table name with dataframe
        df = pd.DataFrame(rest, columns = header)
        mapping[entry] = df
        df.to_csv(f'output/{entry}.csv')
        print(mapping)

    return mapping

# print(read_excel("/workspaces/csv/016-TextFiles/016-MSPTDA-Excel.xlsx")['dCategory'])
